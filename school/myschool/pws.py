import openpyxl as xl

from .models import *


def store():
    wb = xl.load_workbook("C:\\My desktop\\School_project\\school\\myschool\\Student_details.xlsx")
    sheet = wb['Sheet1']
    row = 5
    col = 3
    stu = StudentLogin.objects.all()
    for st in stu:
        curr_cell = sheet.cell(row,col)
        curr_cell.value = st.student_id
        second_cell = sheet.cell(row,col+1)
        second_cell.value = st.login_id
        third_cell = sheet.cell(row,col + 2)
        third_cell.value = st.login_password
        row += 1
    wb.save('student_details1.xlsx')


